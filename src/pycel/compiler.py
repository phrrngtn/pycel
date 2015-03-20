

import collections
import logging

from logging import info, debug

import openpyxl

from openpyxl.exceptions import (
    CellCoordinatesException,
)

from .util import (
    flatten,
    is_range,
    resolve_range,
    split_address,
    split_range,
    uniqueify,
)

from .ast import (
    build_ast,
    RangeNode,
)

from .tokenizer import (
    ExcelParser,
    f_token,
    shunting_yard
)



# XXX: definitely needs a definition/spec
def destination2local(d):
    return d[0][d[1]].coordinate

def apply_names(e, m):
    '''process the stream of tokens from the Excel formula parser and rewrite
    any cell-references as a named range if the reference is present
    in map, m
    '''
    for t in e:
        if isinstance(t,RangeNode) and t.tvalue in m:
            t.tvalue = m[t.tvalue]
        yield t


class ExcelCompiler(object):
    def __init__(self, path=None, workbook=None):
        if path is None:
            self.wb = workbook
        else:
            self.wb = openpyxl.load_workbook(path)

    #XXX this will break if the same cell has multiple names associated with it.
    @property
    def name_reverse_map(self):
        return dict(map(lambda x: (destination2local(x.destinations[0]), x.name), self.wb.get_named_ranges()))

    @property
    def name_map(self):
        return dict(map(lambda x: (x.name, destination2local(x.destinations[0])), self.wb.get_named_ranges()))

    def named_range2code(self, name):
        wb = self.wb
        nr = wb.get_named_range(name)
        (ws, seed) = nr.destinations[0]
        seed_cell = ws.cell(coordinate=seed)

        tokens = self.parse(seed_cell, self.name_map)
        named = apply_names(tokens, self.name_reverse_map)
        G,root = build_ast(named)
        return root.emit(G, context=None)

    def parse(self, cell, name_map):
        '''demonstrate recursive descent of an Excel formula.  'flatten' the
        formula expression tree as a RPN stream of tokens.
        '''
        ws = cell.parent
        wb = ws.parent

        debug("parsing %s data_type %s", cell.value, cell.data_type)
        if cell.data_type == 'f':
            e = shunting_yard(cell.value)
        else:
            e = shunting_yard(str(cell.value))

        for tok in e:
            if isinstance(tok,RangeNode):
                cell_reference = tok.tvalue
                try:
                    dep = ws.cell(coordinate=cell_reference)
                except CellCoordinatesException:
                    dep = ws.cell(coordinate=name_map[cell_reference])

                debug("dependency '%s' type=%s value='%s'", tok.tvalue, dep.data_type, dep.value)
                if dep.data_type == 'f':
                    for tok1 in self.parse(dep, name_map):
                        yield tok1
                    continue
            yield tok

